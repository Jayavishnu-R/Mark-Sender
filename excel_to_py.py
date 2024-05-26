from openpyxl import load_workbook
import os


class Student:

    def __init__(self, student_id, name, ph):
        self.student_id = student_id
        self.name = name
        self.ph = ph


class Mark:
    def __init__(self, student_id, score):
        self.student_id = student_id
        self.score = score


def student_d():
    workboo = load_workbook("details.xlsx")
    students = []
    sheet = workboo.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id, name, ph = row
        student = Student(student_id, name, ph)

        students.append(student)
    return students


def process_excel_file(file_path):

    marks = []
    workbook = load_workbook(file_path)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=True):

        student_id, name, score = row
        mark = Mark(student_id, score)

        marks.append(mark)
    return marks


# Usage example
def main():
    folder_path = "./excel/"
    msgtemp = {}
    students = student_d()
    for student in students:
        # print(f"Student ID: {student.student_id}, Name: {student.name},ph:{student.ph}")
        msgtemp[student.student_id] = [student.name, student.ph, {"marks": {}}]

    # print(msgtemp)

    files = os.listdir(folder_path)
    # print(files)

    for file_name in files:

        mark_data = process_excel_file(folder_path + file_name)
        data = {}

        for mark in mark_data:
            msgtemp[mark.student_id][2]["marks"][file_name[:-5]] = mark.score

        #  print(
        #     f"Student ID: {mark.student_id}, {file_name[:-5]} , Score: {mark.score}"
        # )

    # print(msgtemp)
    msg1 = {}
    for keys, i in msgtemp.items():

        msg = ""
        for sub, j in i[2]["marks"].items():
            msg += f"\n{sub}: {j} marks"

        msg1[i[1]] = (
            "Warm Greeting!!! your son/daughter "
            + i[0]
            + " has secured the following marks in the 1st internal assessment:\n"
            + msg
            + "\n\nBest Regards,\nDr. C.P Indhumathi,\nClass Co-ordinator,\nIII Year CSE,\nAnna University."
        )
        print(msg1[i[1]])
    msgtemp.clear()
    # print(msg1)
    return msg1


if __name__ == "__main__":
    main()
